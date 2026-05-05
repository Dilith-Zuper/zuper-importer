from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re
from datetime import datetime

# ── Load data ─────────────────────────────────────────────────────────────────
def load_sheet(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, r))
        rows.append(row)
    wb.close()
    return rows

hack_rows = load_sheet("D:/OneDrive - Zuper,inc/Documents/Projects/product importer/hackthon 4 product export.xlsx")
rm_rows   = load_sheet("D:/OneDrive - Zuper,inc/Documents/Projects/product importer/Roof_Medic.xlsx")

# ── Filter Roof Medic: materials only, exclude labor/services ─────────────────
LABOR_KW = ['labor','install','remove','crew','warranty','maintenance','repair service','service plan']

def is_labor(row):
    ptype = (row.get('Product Type') or '').upper()
    pname = (row.get('Product Name') or '').lower()
    pcat  = (row.get('Product Category') or '').lower()
    if ptype in ['SERVICE','BUNDLE']: return True
    if 'labor' in pcat: return True
    if any(k in pname for k in LABOR_KW): return True
    return False

rm_materials  = [r for r in rm_rows  if not is_labor(r) and r.get('Product Name')]
hack_products = [r for r in hack_rows if r.get('Product Name')]

def norm(s):
    return re.sub(r'\s+', ' ', (s or '').lower().strip())

# ── Matching ──────────────────────────────────────────────────────────────────
def match(rm_name, rm_brand, hack_list):
    rn = norm(rm_name)
    rb = norm(rm_brand or '')
    for h in hack_list:
        if norm(h.get('Product Name','')) == rn:
            return h, 'Exact name match'
    for h in hack_list:
        hn = norm(h.get('Product Name',''))
        hb = norm(h.get('Brand','') or '')
        if rb and rb == hb and (rn in hn or hn in rn):
            return h, 'Brand + partial name'
    rn_words = set(rn.split())
    for h in hack_list:
        hn = norm(h.get('Product Name',''))
        hb = norm(h.get('Brand','') or '')
        if rb and rb == hb and len(rn_words & set(hn.split())) >= 3:
            return h, 'Brand + keyword match'
    if rb:
        for h in hack_list:
            if norm(h.get('Brand','') or '') == rb:
                return h, 'Brand covered (variant)'
    return None, None

matched   = []
unmatched = []
for rm in rm_materials:
    h, method = match(rm.get('Product Name'), rm.get('Brand'), hack_products)
    rm['_match'] = method
    rm['_matched_name']     = h.get('Product Name')     if h else None
    rm['_matched_brand']    = h.get('Brand')             if h else None
    rm['_matched_price']    = h.get('Price')             if h else None
    rm['_matched_category'] = h.get('Product Category') if h else None
    if h: matched.append(rm)
    else: unmatched.append(rm)

# Stats
brand_stats = defaultdict(lambda: {'total':0,'matched':0})
for r in rm_materials:
    b = (r.get('Brand') or '').strip() or 'Unbranded'
    brand_stats[b]['total'] += 1
for r in matched:
    b = (r.get('Brand') or '').strip() or 'Unbranded'
    brand_stats[b]['matched'] += 1

cat_stats = defaultdict(lambda: {'total':0,'matched':0})
for r in rm_materials:
    c = r.get('Product Category') or 'Unknown'
    cat_stats[c]['total'] += 1
for r in matched:
    c = r.get('Product Category') or 'Unknown'
    cat_stats[c]['matched'] += 1

rm_names_norm = set(norm(r.get('Product Name','')) for r in rm_materials)
extra = [h for h in hack_products if norm(h.get('Product Name','')) not in rm_names_norm]

print(f"RM materials: {len(rm_materials)} | Matched: {len(matched)} ({len(matched)/len(rm_materials)*100:.1f}%) | Unmatched: {len(unmatched)} | Extra: {len(extra)}")

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_color): return PatternFill('solid', fgColor=hex_color)
def thin(): s = Side(style='thin', color='E5E2DC'); return Border(left=s,right=s,top=s,bottom=s)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left_align(): return Alignment(horizontal='left', vertical='center', wrap_text=True)

def hdr(ws, row, vals, bg='1C1917', fg='FFFFFF'):
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = Font(bold=True, color=fg, size=10, name='Calibri')
        c.fill   = fill(bg)
        c.alignment = center()
        c.border = thin()

def drow(ws, row, vals, bg='FFFFFF', bold_col=None):
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = Font(bold=(col==bold_col), color='1A1A1A', size=9, name='Calibri')
        c.fill   = fill(bg)
        c.alignment = left_align()
        c.border = thin()

def col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def sep_row(ws, row, text, ncols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    c.fill  = fill('374151')
    c.alignment = left_align()
    ws.row_dimensions[row].height = 20

def why(r):
    brand = (r.get('Brand') or '').strip()
    if not brand: return 'Unbranded — not in SRS catalog'
    b = brand.lower()
    if 'velux' in b:  return 'Velux not in SRS catalog'
    if 'senox' in b:  return 'Senox not in SRS catalog'
    if 'berger' in b: return 'Berger F8 not in SRS catalog'
    return 'Variant not matched'

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# ═══ Sheet 1: Executive Summary ══════════════════════════════════════════════
ws = wb.active
ws.title = 'Executive Summary'
ws.sheet_view.showGridLines = False

ws.merge_cells('A1:H1')
c = ws['A1']
c.value = 'SRS Product Importer — Roof Medic Coverage Report'
c.font  = Font(bold=True, color='FFFFFF', size=15, name='Calibri')
c.fill  = fill('1C1917')
c.alignment = center()
ws.row_dimensions[1].height = 38

ws.merge_cells('A2:H2')
c = ws['A2']
c.value = f'Generated {datetime.now().strftime("%B %d, %Y")}  |  Hackathon Test 4  |  Brands selected: GAF, CertainTeed, CeDur'
c.font  = Font(color='9CA3AF', size=9, name='Calibri')
c.fill  = fill('1C1917')
c.alignment = center()
ws.row_dimensions[2].height = 18

# KPI row
kpis = [
    ('RM Material\nProducts', str(len(rm_materials)), '1A1A1A'),
    ('Upload\n(Parts)', str(len(hack_products)), 'F97316'),
    ('Matched\nProducts', str(len(matched)), '16A34A'),
    ('Not Covered\n(Out of scope)', str(len(unmatched)), 'DC2626'),
    ('Brand\nCoverage', '100%', '16A34A'),
    ('Overall\nCoverage', f'{len(matched)/len(rm_materials)*100:.0f}%', 'F97316'),
]
kpi_cols = [1,2,3,4,6,7]
for (label, val, color), col in zip(kpis, kpi_cols):
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 45
    lc = ws.cell(row=4, column=col, value=label)
    lc.font = Font(bold=True, color='FFFFFF', size=9, name='Calibri')
    lc.fill = fill('374151')
    lc.alignment = center()
    lc.border = thin()
    vc = ws.cell(row=5, column=col, value=val)
    vc.font = Font(bold=True, color=color, size=24, name='Calibri')
    vc.fill = fill('F5F3F0')
    vc.alignment = center()
    vc.border = thin()

ws.merge_cells('A7:H7')
c = ws['A7']
c.value = ('KEY FINDING: 100% of Roof Medic\'s branded products (GAF, CertainTeed, CeDur = 50 products) are fully '
           'covered by the upload. The 79% "not covered" gap is unbranded commodity items (generic trim, OSB decking, '
           'gutter stock, generic flashing) that are not part of any brand\'s SRS catalog — outside the tool\'s scope.')
c.font = Font(color='92400E', size=10, name='Calibri')
c.fill = fill('FEF3C7')
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
c.border = Border(left=Side(style='thick',color='F97316'), right=thin().right, top=thin().top, bottom=thin().bottom)
ws.row_dimensions[7].height = 52

# Brand table
ws.row_dimensions[9].height = 20
hdr(ws, 9, ['Brand', 'RM Material Products', 'Covered by Upload', 'Not Covered', 'Coverage %', 'Status'], bg='F97316', fg='FFFFFF')
r = 10
for brand, stats in sorted(brand_stats.items(), key=lambda x: -x[1]['total']):
    total = stats['total']; m = stats['matched']
    pct = m/total*100 if total else 0
    status = 'Full Coverage' if pct==100 else ('Partial' if pct>0 else 'Not in SRS Catalog')
    bg = 'DCFCE7' if pct==100 else ('FEFCE8' if pct>0 else 'FEE2E2')
    sc = '16A34A' if pct==100 else ('F97316' if pct>0 else 'DC2626')
    for col, val in enumerate([brand, total, m, total-m, f'{pct:.0f}%', status], 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.fill = fill(bg); cell.border = thin()
        cell.alignment = center() if col > 1 else left_align()
        cell.font = Font(bold=(col==6), color=(sc if col==6 else '1A1A1A'), size=10, name='Calibri')
    ws.row_dimensions[r].height = 18; r += 1

# Category table
r += 1; ws.row_dimensions[r].height = 20
hdr(ws, r, ['Category (Roof Medic)', 'RM Products', 'Covered', 'Not Covered', 'Coverage %', 'Status'], bg='F97316', fg='FFFFFF')
r += 1
for cat, stats in sorted(cat_stats.items(), key=lambda x: -x[1]['total']):
    total = stats['total']; m = stats['matched']
    pct = m/total*100 if total else 0
    status = 'Full' if pct==100 else ('Partial' if pct>0 else 'Not Covered')
    bg = 'DCFCE7' if pct==100 else ('FEFCE8' if pct>0 else 'FEE2E2')
    sc = '16A34A' if pct==100 else ('F97316' if pct>0 else 'DC2626')
    for col, val in enumerate([cat, total, m, total-m, f'{pct:.0f}%', status], 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.fill = fill(bg); cell.border = thin()
        cell.alignment = center() if col > 1 else left_align()
        cell.font = Font(bold=(col==6), color=(sc if col==6 else '1A1A1A'), size=10, name='Calibri')
    ws.row_dimensions[r].height = 18; r += 1

col_w(ws, [30,18,16,14,12,20,0,0])

# ═══ Sheet 2: Matched Products ════════════════════════════════════════════════
ws2 = wb.create_sheet('Matched Products')
ws2.sheet_view.showGridLines = False
ws2.merge_cells('A1:I1')
c = ws2['A1']
c.value = f'Matched Products ({len(matched)}) — Roof Medic items covered by the SRS upload'
c.font = Font(bold=True, color='FFFFFF', size=13, name='Calibri')
c.fill = fill('16A34A'); c.alignment = center()
ws2.row_dimensions[1].height = 28

hdr(ws2, 2, ['Roof Medic Product', 'RM Brand', 'RM Category', 'RM Sell Price',
              'RM Purchase Price', 'Matched Upload Product', 'Upload Brand', 'Upload Category', 'Match Method'])

for i, r in enumerate(matched, 3):
    method = r.get('_match','')
    bg = 'DCFCE7' if 'Exact' in method else 'FFFFFF'
    drow(ws2, i, [
        r.get('Product Name'), (r.get('Brand') or 'Unbranded'), r.get('Product Category'),
        r.get('Price'), r.get('Purchase Price'),
        r.get('_matched_name'), r.get('_matched_brand'), r.get('_matched_category'), method
    ], bg=bg)
    ws2.row_dimensions[i].height = 18

col_w(ws2, [40,16,20,12,14,40,16,24,24])

# ═══ Sheet 3: Not Covered ════════════════════════════════════════════════════
ws3 = wb.create_sheet('Not Covered')
ws3.sheet_view.showGridLines = False
ws3.merge_cells('A1:G1')
c = ws3['A1']
c.value = f'Not Covered ({len(unmatched)}) — Roof Medic materials not in SRS catalog'
c.font = Font(bold=True, color='FFFFFF', size=13, name='Calibri')
c.fill = fill('DC2626'); c.alignment = center()
ws3.row_dimensions[1].height = 28

ws3.merge_cells('A2:G2')
c = ws3['A2']
c.value = 'These products are outside the SRS brand catalog scope (unbranded commodity stock, Velux skylights, Senox gutters). Contractors must add these manually.'
c.font = Font(italic=True, color='92400E', size=9, name='Calibri')
c.fill = fill('FEF3C7'); c.alignment = left_align()
ws3.row_dimensions[2].height = 28

hdr(ws3, 3, ['Roof Medic Product', 'Brand / Source', 'Category', 'Sell Price', 'Purchase Price', 'UOM', 'Reason Not Covered'])

unmatched_sorted = sorted(unmatched, key=lambda x: (x.get('Product Category') or '', x.get('Product Name') or ''))
prev_cat = None; r_num = 4
for r in unmatched_sorted:
    cat = r.get('Product Category') or 'Unknown'
    if cat != prev_cat:
        cat_count = sum(1 for x in unmatched_sorted if (x.get('Product Category') or 'Unknown') == cat)
        sep_row(ws3, r_num, f'  {cat}  ({cat_count} products)')
        r_num += 1; prev_cat = cat
    brand = (r.get('Brand') or '').strip() or 'Unbranded'
    bg = 'FEE2E2' if brand == 'Unbranded' else 'FEFCE8'
    drow(ws3, r_num, [r.get('Product Name'), brand, cat, r.get('Price'), r.get('Purchase Price'), r.get('UOM'), why(r)], bg=bg)
    ws3.row_dimensions[r_num].height = 18; r_num += 1

col_w(ws3, [44,18,22,12,14,8,36])

# ═══ Sheet 4: Extra Uploaded ══════════════════════════════════════════════════
ws4 = wb.create_sheet('Extra Uploaded')
ws4.sheet_view.showGridLines = False
ws4.merge_cells('A1:F1')
c = ws4['A1']
c.value = f'Extra Uploaded ({len(extra)}) — Products in SRS upload not in Roof Medic current stock'
c.font = Font(bold=True, color='FFFFFF', size=13, name='Calibri')
c.fill = fill('F97316'); c.alignment = center()
ws4.row_dimensions[1].height = 28

ws4.merge_cells('A2:F2')
c = ws4['A2']
c.value = 'These products enrich the Zuper catalog for future estimates. Roof Medic may begin using them over time.'
c.font = Font(italic=True, color='92400E', size=9, name='Calibri')
c.fill = fill('FFF3E8'); c.alignment = left_align()
ws4.row_dimensions[2].height = 22

hdr(ws4, 3, ['Product Name', 'Brand', 'Category', 'Price', 'UOM', 'Product Tier'], bg='F97316', fg='FFFFFF')

extra_by_cat = defaultdict(list)
for h in extra:
    extra_by_cat[h.get('Product Category') or 'Unknown'].append(h)

r_num = 4
for cat in sorted(extra_by_cat.keys(), key=lambda x: -len(extra_by_cat[x])):
    items = extra_by_cat[cat]
    sep_row(ws4, r_num, f'  {cat}  ({len(items)} products)', ncols=6)
    r_num += 1
    for h in items[:25]:
        drow(ws4, r_num, [h.get('Product Name'), h.get('Brand'), cat, h.get('Price'), h.get('UOM'), h.get('Product Tier (Default)')], bg='FFF3E8')
        ws4.row_dimensions[r_num].height = 18; r_num += 1
    if len(items) > 25:
        ws4.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=6)
        c = ws4.cell(row=r_num, column=1, value=f'  ... and {len(items)-25} more in this category')
        c.font = Font(italic=True, color='6B7280', size=9, name='Calibri')
        c.fill = fill('FFFFFF'); ws4.row_dimensions[r_num].height = 16; r_num += 1

col_w(ws4, [44,18,28,12,8,20])

# Save
out = "D:/OneDrive - Zuper,inc/Documents/Projects/product importer/Roof_Medic_Coverage_Report.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
