import sys, io, csv, collections, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = 'D:/OneDrive - Zuper,inc/Documents/Projects/product importer'

# ── Our 19 standard tokens ────────────────────────────────────────────────────
STANDARD_TOKENS = {
    'Total Roof Area',
    'Total Eaves Length',
    'Total Rakes Length',
    'Total Ridges Length',
    'Total Hip Length',
    'Total Valleys Length',
    'Total Step Flashing Length',
    'Headwall Flashing',
    'No of Downspouts',
    'No of End Caps',
    'No of Outside Miters',
    'No of Inside Miters',
    'Downspout Elbows',
    'No of Inner Elbows',
    'No of Outer Elbows',
    'Gutter Length',
    'Suggested Waste Percentage %',
    'Total Siding Area',
    'No of Downspouts',
}

# Known naming variants of standard tokens
NAMING_VARIANTS = {
    'Total Eaves':      'Total Eaves Length',
    'Total Rakes':      'Total Rakes Length',
    'Total Valleys':    'Total Valleys Length',
    'Total Hips':       'Total Hip Length',
    'Total Ridge':      'Total Ridges Length',
    'Total Ridge Length': 'Total Ridges Length',
    'Waste Percentage': 'Suggested Waste Percentage %',
    'Suggested Waste Percentage': 'Suggested Waste Percentage %',
    'Waste %':          'Suggested Waste Percentage %',
    'Roof Area':        'Total Roof Area',
    'Eaves Length':     'Total Eaves Length',
    'Rakes Length':     'Total Rakes Length',
    'Hip Length':       'Total Hip Length',
    'Valley Length':    'Total Valleys Length',
    'Ridges Length':    'Total Ridges Length',
    'Step Flashing Length': 'Total Step Flashing Length',
    'Siding Area':      'Total Siding Area',
}

SLOPE_KEYWORDS = ['slope', 'pitch', 'low slope', 'standard slope', 'steep slope',
                  'flat slope', 'roof pitch', 'clerestory', '0/12', '1/12', '2/12',
                  '3/12', '4/12', '5/12', '6/12', '7/12', '8/12', '9/12', '10/12',
                  '11/12', '12/12', '>12']

# ── Styles ────────────────────────────────────────────────────────────────────
DARK    = '1A1A1A'
ORANGE  = 'F97316'
WHITE   = 'FFFFFF'
LGRAY   = 'F5F4F1'
GREEN   = '16A34A'
AMBER   = 'D97706'
RED_C   = 'DC2626'
BLUE    = '2563EB'
PURPLE  = '7C3AED'

def hdr_fill(hex_color): return PatternFill('solid', fgColor=hex_color)
def thin_border():
    s = Side(style='thin', color='E5E2DC')
    return Border(left=s, right=s, top=s, bottom=s)

def write_row(ws, row_idx, values, bold=False, bg=None, fg=DARK, sizes=None, aligns=None, height=16):
    for ci, val in enumerate(values, 1):
        c = ws.cell(row_idx, ci, val)
        c.font = Font(bold=bold, color=fg, size=(sizes[ci-1] if sizes else 10))
        if bg: c.fill = hdr_fill(bg)
        c.border = thin_border()
        al = aligns[ci-1] if aligns else 'left'
        c.alignment = Alignment(horizontal=al, vertical='center', wrap_text=False)
    ws.row_dimensions[row_idx].height = height

def title_row(ws, text, ncols, bg=DARK, fg=WHITE, size=13, height=28):
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    c = ws['A1']
    c.value = text
    c.font = Font(bold=True, size=size, color=fg)
    c.fill = hdr_fill(bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = height
    ws.sheet_view.showGridLines = False

# ── Load data ─────────────────────────────────────────────────────────────────
print('Loading CSVs...')

tokens_rows = []
for fname in ['us_east_zuper_service.measurement_categories.csv',
              'us_west_zuper_service.measurement_categories.csv']:
    with open(f'{BASE}/{fname}', encoding='utf-8') as f:
        for r in csv.DictReader(f):
            tokens_rows.append(r)

formula_rows = []
for fname in ['us_east_zuper_service.cpq_formula.csv',
              'us_west_zuper_service.cpq_formula.csv']:
    with open(f'{BASE}/{fname}', encoding='utf-8') as f:
        for r in csv.DictReader(f):
            formula_rows.append(r)

all_companies = set(r['company_id'] for r in tokens_rows)
TOTAL_COMPANIES = len(all_companies)
print(f'  {len(tokens_rows)} token rows, {len(formula_rows)} formula rows, {TOTAL_COMPANIES} companies')

# ── Build token → company set ─────────────────────────────────────────────────
token_companies: dict[str, set] = collections.defaultdict(set)
for r in tokens_rows:
    tok = r.get('measurement_token_name', '').strip()
    if tok:
        token_companies[tok].add(r['company_id'])

# ── Build formula index ───────────────────────────────────────────────────────
# formula_expr → list of (company_id, formula_name)
formula_by_token: dict[str, list] = collections.defaultdict(list)
for r in formula_rows:
    expr = r.get('formula', '') or ''
    name = r.get('formula_name', '') or ''
    cid  = r.get('company_id', '')
    # extract token-like words (capitalized phrases) from expression
    # also just store the full expression for searching
    formula_by_token['__all__'].append((cid, name, expr))

def count_formulas_using_token(token_name: str) -> int:
    tok_lower = token_name.lower()
    return sum(1 for _, _, expr in formula_by_token['__all__'] if tok_lower in expr.lower())

def get_formula_examples(token_name: str, limit=3) -> list[str]:
    tok_lower = token_name.lower()
    examples = []
    seen = set()
    for _, name, expr in formula_by_token['__all__']:
        if tok_lower in expr.lower() and expr not in seen:
            examples.append(f'{name}: {expr[:80]}')
            seen.add(expr)
            if len(examples) >= limit:
                break
    return examples

# ── Classify tokens ───────────────────────────────────────────────────────────
def classify(name: str, count: int) -> str:
    if name in STANDARD_TOKENS:        return 'Standard'
    if name in NAMING_VARIANTS:        return 'Naming Variant'
    nl = name.lower()
    if any(k in nl for k in SLOPE_KEYWORDS): return 'Slope / Pitch'
    if count >= 50:                    return 'Quasi-Standard'
    if count >= 10:                    return 'Common'
    if count >= 3:                     return 'Limited'
    return 'Niche'

CLASS_COLOR = {
    'Standard':       GREEN,
    'Naming Variant': AMBER,
    'Quasi-Standard': BLUE,
    'Slope / Pitch':  PURPLE,
    'Common':         '0891B2',
    'Limited':        '6B7280',
    'Niche':          '9CA3AF',
}

# ── Sort all tokens ───────────────────────────────────────────────────────────
all_tokens_sorted = sorted(token_companies.items(), key=lambda x: -len(x[1]))

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Token Prevalence (all 802 tokens)
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Token Prevalence'
title_row(ws1, f'All Measurement Tokens — {len(all_tokens_sorted)} unique across {TOTAL_COMPANIES} accounts', 7)

for w, col in zip([38, 14, 12, 14, 12, 12, 50], 'ABCDEFG'):
    ws1.column_dimensions[col].width = w

hdrs = ['Token Name', '# Accounts', '% of Accounts', 'Classification', 'Formula Uses', 'In Our Standard?', 'Formula Example']
write_row(ws1, 2, hdrs, bold=True, bg=DARK, fg=WHITE, height=20,
          aligns=['left','center','center','left','center','center','left'])

for ri, (tok, cos) in enumerate(all_tokens_sorted, 3):
    count     = len(cos)
    pct       = f'{round(100*count/TOTAL_COMPANIES)}%'
    cls       = classify(tok, count)
    fml_count = count_formulas_using_token(tok)
    is_std    = 'Yes' if tok in STANDARD_TOKENS else ''
    examples  = get_formula_examples(tok, 1)
    ex_str    = examples[0][:80] if examples else ''
    bg = WHITE if ri % 2 == 0 else LGRAY
    write_row(ws1, ri, [tok, count, pct, cls, fml_count or '', is_std, ex_str],
              bg=bg, aligns=['left','center','center','left','center','center','left'])
    cls_color = CLASS_COLOR.get(cls, DARK)
    ws1.cell(ri, 4).font = Font(bold=True, color=cls_color, size=10)
    if is_std:
        ws1.cell(ri, 6).font = Font(bold=True, color=GREEN, size=10)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Gap Analysis (our 19 standard tokens)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Our Standard Tokens')
title_row(ws2, 'Gap Analysis — Our 19 Standard Tokens vs. Account Reality', 5, bg=GREEN)
for w, col in zip([38, 14, 12, 10, 40], 'ABCDE'):
    ws2.column_dimensions[col].width = w

write_row(ws2, 2, ['Token Name', '# Accounts', '% of Accounts', 'Risk', 'Missing from (company IDs)'],
          bold=True, bg=GREEN, fg=WHITE, height=20,
          aligns=['left','center','center','center','left'])

for ri, tok in enumerate(sorted(STANDARD_TOKENS), 3):
    cos   = token_companies.get(tok, set())
    count = len(cos)
    pct   = f'{round(100*count/TOTAL_COMPANIES)}%'
    missing = all_companies - cos
    if count == TOTAL_COMPANIES:
        risk = 'None'
        rfg  = GREEN
    elif count >= TOTAL_COMPANIES - 2:
        risk = 'Low'
        rfg  = AMBER
    else:
        risk = 'Medium'
        rfg  = RED_C
    missing_str = ', '.join(sorted(missing)[:5]) + (f' (+{len(missing)-5} more)' if len(missing) > 5 else '') if missing else '—'
    bg = WHITE if ri % 2 == 0 else LGRAY
    write_row(ws2, ri, [tok, count, pct, risk, missing_str],
              bg=bg, aligns=['left','center','center','center','left'])
    ws2.cell(ri, 4).font = Font(bold=True, color=rfg, size=10)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Quasi-Standard Tokens (50+ accounts, not in our list)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Quasi-Standard Tokens')
title_row(ws3, 'Widely-Deployed Tokens Not in Our Standard List (50+ accounts)', 6, bg=BLUE)
for w, col in zip([40, 14, 12, 14, 12, 50], 'ABCDEF'):
    ws3.column_dimensions[col].width = w

write_row(ws3, 2, ['Token Name', '# Accounts', '% Accounts', 'Formula Uses', 'Add to Our List?', 'Formula Example'],
          bold=True, bg=BLUE, fg=WHITE, height=20,
          aligns=['left','center','center','center','center','left'])

quasi = [(tok, cos) for tok, cos in all_tokens_sorted
         if tok not in STANDARD_TOKENS and len(cos) >= 50]

for ri, (tok, cos) in enumerate(quasi, 3):
    count     = len(cos)
    pct       = f'{round(100*count/TOTAL_COMPANIES)}%'
    fml_count = count_formulas_using_token(tok)
    cls       = classify(tok, count)
    recommend = 'Yes — slope service' if cls == 'Slope / Pitch' else ('Consider' if fml_count > 50 else 'No')
    examples  = get_formula_examples(tok, 1)
    ex_str    = examples[0][:80] if examples else ''
    bg = WHITE if ri % 2 == 0 else LGRAY
    write_row(ws3, ri, [tok, count, pct, fml_count or 0, recommend, ex_str],
              bg=bg, aligns=['left','center','center','center','center','left'])
    rec_color = GREEN if recommend.startswith('Yes') else (AMBER if recommend == 'Consider' else '9CA3AF')
    ws3.cell(ri, 5).font = Font(bold=True, color=rec_color, size=10)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Naming Variants
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Naming Variants')
title_row(ws4, 'Naming Variants — Same Meaning as Standard Tokens, Different Name', 6, bg=AMBER, fg=DARK)
for w, col in zip([32, 32, 14, 14, 10, 50], 'ABCDEF'):
    ws4.column_dimensions[col].width = w

write_row(ws4, 2, ['Variant Token Name', 'Maps To (Standard)', '# Accounts (variant)', '# Formulas Using Variant', 'Risk if Missing', 'Formula Example'],
          bold=True, bg=AMBER, fg=DARK, height=20,
          aligns=['left','left','center','center','center','left'])

# Find tokens that look like variants (either in our NAMING_VARIANTS dict or by fuzzy match)
variant_rows = []
for tok, standard in NAMING_VARIANTS.items():
    cos   = token_companies.get(tok, set())
    count = len(cos)
    fml   = count_formulas_using_token(tok)
    ex    = get_formula_examples(tok, 1)
    # Check if any account HAS the variant but NOT the standard
    std_cos  = token_companies.get(standard, set())
    gap_cos  = cos - std_cos
    risk     = f'HIGH — {len(gap_cos)} accounts only have variant' if gap_cos else 'Low — standard also present'
    variant_rows.append((tok, standard, count, fml, risk, ex[0][:80] if ex else ''))

variant_rows.sort(key=lambda x: -x[3])

for ri, (tok, std, count, fml, risk, ex) in enumerate(variant_rows, 3):
    bg = WHITE if ri % 2 == 0 else LGRAY
    write_row(ws4, ri, [tok, std, count, fml, risk, ex],
              bg=bg, aligns=['left','left','center','center','left','left'])
    risk_color = RED_C if risk.startswith('HIGH') else GREEN
    ws4.cell(ri, 5).font = Font(bold=True, color=risk_color, size=10)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Slope Tokens
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Slope Tokens')
title_row(ws5, 'Slope / Pitch Tokens — Cross-Reference with Our Slope-Based Services', 5, bg=PURPLE)
for w, col in zip([42, 14, 12, 14, 50], 'ABCDE'):
    ws5.column_dimensions[col].width = w

write_row(ws5, 2, ['Token Name', '# Accounts', '% Accounts', 'Formula Uses', 'Our Slope Tier Mapping'],
          bold=True, bg=PURPLE, fg=WHITE, height=20,
          aligns=['left','center','center','center','left'])

OUR_TIERS = {
    'low slope': 'Low (3/12-6/12) — $107 tear-off + $30 install',
    'standard slope': 'Standard (7/12-9/12) — $185 tear-off + $77 install',
    'steep slope': 'Steep (10/12-12/12) — $245 tear-off + $107 install',
    'very steep': 'Very Steep (13/12+) — $321 tear-off + $132 install',
    'flat slope': 'Below our Low tier (< 3/12)',
    '2/12': 'Below Low tier',
    '3/12': 'Maps to Low tier boundary',
    '4/12': 'Low tier (3/12-6/12)',
    '5/12': 'Low tier (3/12-6/12)',
    '6/12': 'Low tier (3/12-6/12)',
    '7/12': 'Standard tier (7/12-9/12)',
    '8/12': 'Standard tier (7/12-9/12)',
    '9/12': 'Standard tier (7/12-9/12)',
    '10/12': 'Steep tier (10/12-12/12)',
    '11/12': 'Steep tier (10/12-12/12)',
    '12/12': 'Steep tier (10/12-12/12)',
    '>12': 'Very Steep tier (13/12+)',
}

slope_tokens = [(tok, cos) for tok, cos in all_tokens_sorted
                if any(k in tok.lower() for k in SLOPE_KEYWORDS)]

for ri, (tok, cos) in enumerate(slope_tokens, 3):
    count = len(cos)
    pct   = f'{round(100*count/TOTAL_COMPANIES)}%'
    fml   = count_formulas_using_token(tok)
    tok_l = tok.lower()
    mapping = next((v for k, v in OUR_TIERS.items() if k in tok_l), 'No direct mapping')
    bg = WHITE if ri % 2 == 0 else LGRAY
    write_row(ws5, ri, [tok, count, pct, fml or 0, mapping],
              bg=bg, aligns=['left','center','center','center','left'])
    map_color = GREEN if 'tier' in mapping.lower() else (AMBER if 'boundary' in mapping.lower() else '9CA3AF')
    ws5.cell(ri, 5).font = Font(color=map_color, size=10)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Custom Tokens Actually in Formulas
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Custom Tokens in Formulas')
title_row(ws6, 'Non-Standard Tokens Actually Referenced in CPQ Formulas', 6, bg=RED_C)
for w, col in zip([40, 14, 12, 14, 20, 60], 'ABCDEF'):
    ws6.column_dimensions[col].width = w

write_row(ws6, 2, ['Token Name', '# Accounts', '% Accounts', 'Formula Uses', 'Verdict', 'Formula Example'],
          bold=True, bg=RED_C, fg=WHITE, height=20,
          aligns=['left','center','center','center','left','left'])

custom_in_formulas = []
for tok, cos in all_tokens_sorted:
    if tok in STANDARD_TOKENS: continue
    fml = count_formulas_using_token(tok)
    if fml > 0:
        custom_in_formulas.append((tok, cos, fml))

custom_in_formulas.sort(key=lambda x: -x[2])

for ri, (tok, cos, fml) in enumerate(custom_in_formulas, 3):
    count = len(cos)
    pct   = f'{round(100*count/TOTAL_COMPANIES)}%'
    cls   = classify(tok, count)
    ex    = get_formula_examples(tok, 1)
    ex_str = ex[0][:100] if ex else ''
    if tok in NAMING_VARIANTS:
        verdict = f'Variant of "{NAMING_VARIANTS[tok]}" — standard covers it'
        vfg = AMBER
    elif cls == 'Slope / Pitch':
        verdict = 'Slope token — could map to our service tiers'
        vfg = PURPLE
    elif count >= 50:
        verdict = 'Quasi-standard — consider adding'
        vfg = BLUE
    else:
        verdict = 'Account-specific — leave as custom'
        vfg = '6B7280'
    bg = WHITE if ri % 2 == 0 else LGRAY
    write_row(ws6, ri, [tok, count, pct, fml, verdict, ex_str],
              bg=bg, aligns=['left','center','center','center','left','left'])
    ws6.cell(ri, 5).font = Font(bold=True, color=vfg, size=10)

# ── Save ──────────────────────────────────────────────────────────────────────
out = f'{BASE}/Token_Analysis.xlsx'
wb.save(out)
print(f'\nSaved: {out}')
print(f'Sheets: {[s.title for s in wb.worksheets]}')
print(f'\nSummary:')
print(f'  Total unique tokens: {len(all_tokens_sorted)}')
print(f'  Standard tokens present in all accounts: {sum(1 for t in STANDARD_TOKENS if len(token_companies.get(t,set())) == TOTAL_COMPANIES)}')
print(f'  Quasi-standard tokens (50+ accounts, not our list): {len(quasi)}')
print(f'  Slope/pitch tokens: {len(slope_tokens)}')
print(f'  Custom tokens actually used in formulas: {len(custom_in_formulas)}')
