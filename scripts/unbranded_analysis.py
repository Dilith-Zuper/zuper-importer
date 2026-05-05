import sys, io, csv, collections, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = 'D:/OneDrive - Zuper,inc/Documents/Projects/product importer'

# ── Keyword sets per RM category ─────────────────────────────────────────────
CATEGORY_KEYWORDS = {
    'Flashing':               ['flash', 'valley', 'apron', 'step flash', 'counter flash', 'headwall', 'endwall'],
    'Fasteners':              ['nail', 'fastener', 'screw', 'coil nail', 'caulk', 'np1', 'button cap'],
    'Boots and Penetrations': ['boot', 'pipe boot', 'zipper boot', 'storm collar', 'penetration'],
    'Drip Edge':              ['drip edge', 'drip cap', 'eave drip', 'rake edge'],
    'Underlayment':           ['underlayment', 'felt', 'ice and water', 'ice & water', 'synthetic', 'leak barrier', 'moisture guard'],
    'Ventilation':            ['ridge vent', 'box vent', 'soffit vent', 'turbine', 'exhaust vent', 'filtered ridge', 'broan'],
    'Gutters and Downspouts': ['gutter', 'downspout', 'elbow', 'end cap', 'miter', 'gutter guard', 'leaf'],
    'Soffit':                 ['soffit'],
    'Hip and Ridge':          ['hip cap', 'ridge cap', 'hip ridge'],
    'Starter':                ['starter', 'starter strip'],
    'Decking':                ['osb', 'decking', 'plywood', 'sheathing'],
    'Skylights':              ['skylight', 'velux'],
    'Trim':                   ['trim', 'fascia', 'j-trim', 'f-channel', 'corner post'],
    'Roofing Materials':      ['shingle', 'tile', 'slate', 'metal panel', 'cedar shake'],
    'Low Slope':              ['tpo', 'epdm', 'modified bitumen', 'built-up', 'flat roof'],
    'Parts':                  ['ridge', 'cap', 'accessory', 'component'],
    'Siding':                 ['siding', 'lap siding', 'vinyl siding'],
    'CREW':                   [],
    'Custom':                 [],
}

def cat_keywords(category):
    return CATEGORY_KEYWORDS.get(category, [])

def name_keywords(product_name):
    STOP = {'the','and','for','per','with','box','bx','ea','rl','sq','lf','each','roll',
            'bundle','pol','ctn','are','not','our','this','that','from','has','have',
            'been','will','was','its','all','can','but'}
    return [w for w in re.split(r'[\s\-/(),\"\']+', product_name.lower())
            if len(w) >= 4 and w not in STOP and not w.isdigit()]

def matches_keywords(text, keywords):
    t = text.lower()
    return any(k in t for k in keywords)

GENERIC_WORDS = {'metal','flow','style','type','size','standard','product','material',
                 'series','grade','class','model','brand','general','master','value'}

def name_match_score(account_name, rm_name, cat_kws):
    """Count of rm_name distinctive words that appear in account_name.
    Words must also be category-relevant OR not generic.
    Threshold caller should use: score >= max(1, len(kws)//2)
    """
    kws = [w for w in name_keywords(rm_name) if w not in GENERIC_WORDS]
    # Boost: prefer words that are also in category keywords
    distinctive = [w for w in kws if any(w in ck for ck in cat_kws)] or kws
    if not distinctive:
        return 0
    t = account_name.lower()
    return sum(1 for k in distinctive if k in t)

# ── Load RM unbranded items ───────────────────────────────────────────────────
print('Loading Roof Medic gap items...')
wb_rm = openpyxl.load_workbook(f'{BASE}/Roof_Medic_Coverage_Report.xlsx', data_only=True)
ws_rm = wb_rm['Not Covered']

rm_items = []
for row in ws_rm.iter_rows(min_row=4, values_only=True):
    name, brand, cat, sell, buy, uom, reason = (list(row) + [None]*7)[:7]
    if not name or not cat:
        continue
    name = str(name).strip()
    cat = str(cat).strip() if cat else ''
    # Skip section header rows (start with spaces or are all-caps group labels)
    if name.startswith(' ') or (name.startswith('(') and 'products' in name.lower()):
        continue
    if len(name) < 3:
        continue
    sell_val = float(sell) if sell and str(sell).replace('.','').isdigit() else None
    rm_items.append({
        'name': name,
        'brand': str(brand).strip() if brand else 'Unbranded',
        'category': cat,
        'sell_price': sell_val,
        'reason': str(reason).strip() if reason else '',
    })

print(f'  {len(rm_items)} RM gap items loaded')

# ── Load account products (East + West) ──────────────────────────────────────
print('Loading account products...')
account_rows = []
for fname in ['us_east_zuper_service.products.csv', 'us_west_zuper_service.products.csv']:
    with open(f'{BASE}/{fname}', encoding='utf-8') as f:
        for r in csv.DictReader(f):
            if r.get('product_type') in ('PARTS', 'PRODUCT'):
                account_rows.append(r)

all_companies = set(r['company_id'] for r in account_rows)
total_companies = len(all_companies)
print(f'  {len(account_rows)} account product rows, {total_companies} companies')

# ── Load SRS catalog ─────────────────────────────────────────────────────────
print('Loading SRS catalog...')
wb_srs = openpyxl.load_workbook(f'{BASE}/SRS Catalog Export.xlsx', data_only=True)
ws_srs = wb_srs['📦 Products']
srs_headers = [ws_srs.cell(1, c).value for c in range(1, ws_srs.max_column + 1)]
SRS_NAME_IDX = srs_headers.index('Product Name')
SRS_CAT_IDX  = srs_headers.index('Category')
SRS_MFR_IDX  = srs_headers.index('Manufacturer')

srs_products = []
for row in ws_srs.iter_rows(min_row=2, values_only=True):
    name = row[SRS_NAME_IDX]
    cat  = row[SRS_CAT_IDX]
    mfr  = row[SRS_MFR_IDX]
    if name:
        srs_products.append({'name': str(name), 'category': str(cat or ''), 'manufacturer': str(mfr or '')})

print(f'  {len(srs_products)} SRS products loaded')

# ── Per-item analysis ─────────────────────────────────────────────────────────
print('Analysing each item...')
results = []

for item in rm_items:
    cat_kws  = cat_keywords(item['category'])   # curated category keywords only
    item_kws = name_keywords(item['name'])       # distinctive name words

    # Account prevalence — use category keywords (not name) for category-level signal
    cat_companies = set()
    name_companies = set()
    for row in account_rows:
        pname = row.get('product_name', '')
        if cat_kws and matches_keywords(pname, cat_kws):
            cat_companies.add(row['company_id'])
        # Name match: at least 1 category-relevant distinctive word from RM name
        if name_match_score(pname, item['name'], cat_kws) >= 1:
            name_companies.add(row['company_id'])

    n_companies      = len(cat_companies)           # category-level coverage
    n_exact_matches  = len(name_companies)          # specific item-level coverage
    pct = round(100 * n_companies / total_companies, 1)

    # SRS counterparts — use category keywords only (not name) to avoid false positives
    srs_matches = [p for p in srs_products if cat_kws and matches_keywords(p['name'], cat_kws)]
    # Narrow down with at least 1 distinctive name word for better relevance
    srs_name_matches = [p for p in srs_matches if item_kws and matches_keywords(p['name'], item_kws)]
    # Use name matches if any, fall back to category matches for brands
    srs_for_brands = srs_name_matches if srs_name_matches else srs_matches
    has_srs = len(srs_matches) > 0
    mfrs = list(dict.fromkeys(p['manufacturer'] for p in srs_for_brands if p['manufacturer']))[:6]
    top_brands = ', '.join(mfrs) if mfrs else '—'
    srs_count = len(srs_matches)

    # Recommendation
    if n_companies >= 10 and has_srs:
        rec = 'Add to import scope'
    elif n_companies >= 10 and not has_srs:
        rec = 'Source separately'
    elif n_companies >= 5:
        rec = 'Consider adding'
    else:
        rec = 'Low priority'

    results.append({**item, 'n_companies': n_companies, 'n_exact': n_exact_matches,
                    'pct': pct, 'has_srs': has_srs, 'srs_count': srs_count,
                    'top_brands': top_brands, 'srs_matches': srs_matches,
                    'recommendation': rec})

results.sort(key=lambda x: -x['n_companies'])

# ── Category summary ──────────────────────────────────────────────────────────
cat_summary = collections.defaultdict(lambda: {'items': 0, 'companies_list': [], 'srs_total': 0, 'srs_mfrs': set()})
for r in results:
    c = r['category']
    cat_summary[c]['items'] += 1
    cat_summary[c]['companies_list'].append(r['n_companies'])
    cat_summary[c]['srs_total'] += r['srs_count']
    for m in r['top_brands'].split(', '):
        if m != '—':
            cat_summary[c]['srs_mfrs'].add(m)

# ── Build Excel ───────────────────────────────────────────────────────────────
print('Writing Excel...')
wb = openpyxl.Workbook()

# Styles
ORANGE      = 'FF6B35'
DARK        = '1A1A1A'
LIGHT_GRAY  = 'F5F4F1'
GREEN       = '16A34A'
AMBER       = 'D97706'
RED         = 'DC2626'
BLUE        = '2563EB'
WHITE       = 'FFFFFF'

def hdr_font(bold=True, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size)

def cell_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color='E5E2DC')
    return Border(left=s, right=s, top=s, bottom=s)

def write_row(ws, row_idx, values, bold=False, bg=None, fg=DARK, sizes=None, aligns=None, wrap=False):
    for ci, val in enumerate(values, 1):
        c = ws.cell(row_idx, ci, val)
        c.font = Font(bold=bold, color=fg, size=sizes[ci-1] if sizes else 10)
        if bg:
            c.fill = cell_fill(bg)
        c.border = thin_border()
        al = aligns[ci-1] if aligns else 'left'
        c.alignment = Alignment(horizontal=al, vertical='center', wrap_text=wrap)


# ════════════════════════════════════════════════════════════════
# SHEET 1 — Executive Summary
# ════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Executive Summary'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 32
for col in 'BCDEFG':
    ws1.column_dimensions[col].width = 18

# Title
ws1.merge_cells('A1:G1')
t = ws1['A1']
t.value = 'Unbranded Gap Analysis — Account Prevalence & SRS Coverage'
t.font = Font(bold=True, size=16, color=WHITE)
t.fill = cell_fill(DARK)
t.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 36

# KPIs
high_demand   = sum(1 for r in results if r['n_companies'] >= 10)
has_srs_count = sum(1 for r in results if r['has_srs'])
add_scope     = sum(1 for r in results if r['recommendation'] == 'Add to import scope')
kpis = [
    ('Gap Items (RM unbranded)', len(results), DARK),
    ('High-Demand (10+ accts)', high_demand, ORANGE),
    ('Have SRS Equivalent', has_srs_count, GREEN),
    ('Recommended to Add', add_scope, BLUE),
]
ws1.row_dimensions[2].height = 8
ws1.row_dimensions[3].height = 56
ws1.row_dimensions[4].height = 22
for i, (label, val, color) in enumerate(kpis, 1):
    col = get_column_letter(i * 2 - 1)
    ws1.merge_cells(f'{col}3:{get_column_letter(i*2)}3')
    c = ws1[f'{col}3']
    c.value = f'{val}\n{label}'
    c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = cell_fill(color)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Category table
ws1.row_dimensions[5].height = 8
hdr_row = 6
ws1.row_dimensions[hdr_row].height = 22
cat_headers = ['Category', 'RM Items', 'Avg Acct Coverage', '% of 54 Accts', 'SRS Products', 'SRS Brands', 'Verdict']
write_row(ws1, hdr_row, cat_headers, bold=True, bg=DARK, fg=WHITE,
          aligns=['left','center','center','center','center','center','left'])

cats_sorted = sorted(cat_summary.items(), key=lambda x: -max(x[1]['companies_list']) if x[1]['companies_list'] else 0)
for ri, (cat, data) in enumerate(cats_sorted, hdr_row + 1):
    avg_co = round(sum(data['companies_list']) / len(data['companies_list']), 1) if data['companies_list'] else 0
    avg_pct = round(100 * avg_co / total_companies, 1)
    n_mfrs = len(data['srs_mfrs'])
    if avg_co >= 10 and data['srs_total'] > 0:
        verdict = 'Add to import scope'
        vfg = GREEN
    elif avg_co >= 10:
        verdict = 'Source separately'
        vfg = RED
    elif avg_co >= 5:
        verdict = 'Consider adding'
        vfg = AMBER
    else:
        verdict = 'Low priority'
        vfg = '6B7280'
    bg = WHITE if ri % 2 == 0 else LIGHT_GRAY
    ws1.row_dimensions[ri].height = 18
    row_vals = [cat, data['items'], avg_co, f'{avg_pct}%', data['srs_total'], n_mfrs, verdict]
    write_row(ws1, ri, row_vals, bg=bg, aligns=['left','center','center','center','center','center','left'])
    ws1.cell(ri, 7).font = Font(bold=True, color=vfg, size=10)


# ════════════════════════════════════════════════════════════════
# SHEET 2 — Item-Level Analysis
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Item Analysis')
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 24
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 40
ws2.column_dimensions['I'].width = 22

# Title
ws2.merge_cells('A1:I1')
t = ws2['A1']
t.value = 'Item-Level Analysis — 207 Unbranded Gap Items (sorted by category account coverage)'
t.font = Font(bold=True, size=13, color=WHITE)
t.fill = cell_fill(DARK)
t.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

item_hdrs = ['RM Product', 'Category', 'RM Price', 'Accts w/ Category', '% of 54 Accts', 'Direct Name Matches', 'SRS Equiv?', 'Top SRS Brands', 'Recommendation']
write_row(ws2, 2, item_hdrs, bold=True, bg=ORANGE, fg=WHITE,
          aligns=['left','left','center','center','center','center','center','left','left'])
ws2.row_dimensions[2].height = 20

for ri, r in enumerate(results, 3):
    bg = WHITE if ri % 2 == 0 else LIGHT_GRAY
    price_str = f"${r['sell_price']:.2f}" if r['sell_price'] else '—'
    srs_yn = 'Yes' if r['has_srs'] else 'No'
    rec_color = {'Add to import scope': GREEN, 'Source separately': RED,
                 'Consider adding': AMBER, 'Low priority': '9CA3AF'}.get(r['recommendation'], DARK)
    row_vals = [r['name'], r['category'], price_str, r['n_companies'],
                f"{r['pct']}%", r['n_exact'], srs_yn, r['top_brands'], r['recommendation']]
    write_row(ws2, ri, row_vals, bg=bg, aligns=['left','left','center','center','center','center','center','left','left'])
    ws2.cell(ri, 7).font = Font(bold=True, color=GREEN if r['has_srs'] else RED, size=10)
    ws2.cell(ri, 9).font = Font(bold=True, color=rec_color, size=10)
    ws2.row_dimensions[ri].height = 16


# ════════════════════════════════════════════════════════════════
# SHEET 3 — SRS Counterparts Detail
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('SRS Counterparts')
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 52
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 26

ws3.merge_cells('A1:D1')
t = ws3['A1']
t.value = 'SRS Branded Counterparts — what branded options exist for each gap category'
t.font = Font(bold=True, size=13, color=WHITE)
t.fill = cell_fill(DARK)
t.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 28

cur_row = 2
# Group SRS matches by RM category
cat_srs = collections.defaultdict(list)
for r in results:
    for p in r['srs_matches'][:3]:  # up to 3 per RM item
        cat_srs[r['category']].append(p)

for cat in sorted(cat_srs.keys()):
    # Section header
    ws3.merge_cells(f'A{cur_row}:D{cur_row}')
    c = ws3[f'A{cur_row}']
    c.value = cat
    c.font = Font(bold=True, size=11, color=WHITE)
    c.fill = cell_fill(ORANGE)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws3.row_dimensions[cur_row].height = 20
    cur_row += 1

    # Sub-header
    sub_hdrs = ['Category (SRS)', 'SRS Product Name', 'Manufacturer', '']
    write_row(ws3, cur_row, sub_hdrs, bold=True, bg=LIGHT_GRAY, fg=DARK)
    ws3.row_dimensions[cur_row].height = 16
    cur_row += 1

    # Deduplicate by product name
    seen = set()
    shown = 0
    for p in cat_srs[cat]:
        if p['name'] in seen or shown >= 12:
            continue
        seen.add(p['name'])
        bg = WHITE if shown % 2 == 0 else 'FAF9F7'
        write_row(ws3, cur_row, [p['category'], p['name'], p['manufacturer'], ''], bg=bg)
        ws3.row_dimensions[cur_row].height = 15
        cur_row += 1
        shown += 1

    cur_row += 1  # gap between categories


# ════════════════════════════════════════════════════════════════
# SHEET 4 — Add to Scope Shortlist
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Add to Scope Shortlist')
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 42
ws4.column_dimensions['B'].width = 24
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 14
ws4.column_dimensions['F'].width = 40

ws4.merge_cells('A1:F1')
t = ws4['A1']
t.value = 'Action List — Items Recommended to Add to Import Scope'
t.font = Font(bold=True, size=13, color=WHITE)
t.fill = cell_fill(GREEN)
t.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 28

shortlist = [r for r in results if r['recommendation'] == 'Add to import scope']
ws4['A2'].value = f'{len(shortlist)} items are high-demand (10+ accounts) AND have a branded SRS equivalent ready to import.'
ws4['A2'].font = Font(italic=True, color='374151', size=10)
ws4.merge_cells('A2:F2')

sc_hdrs = ['RM Product', 'Category', 'RM Price', 'Accts w/ Category', '% of 54 Accts', 'Top SRS Brands Available']
write_row(ws4, 3, sc_hdrs, bold=True, bg=GREEN, fg=WHITE,
          aligns=['left','left','center','center','center','left'])
ws4.row_dimensions[3].height = 20

for ri, r in enumerate(shortlist, 4):
    bg = WHITE if ri % 2 == 0 else 'F0FDF4'
    price_str = f"${r['sell_price']:.2f}" if r['sell_price'] else '—'
    write_row(ws4, ri, [r['name'], r['category'], price_str,
                        r['n_companies'], f"{r['pct']}%", r['top_brands']],
              bg=bg, aligns=['left','left','center','center','center','left'])
    ws4.row_dimensions[ri].height = 16


# ── Save ──────────────────────────────────────────────────────────────────────
out = f'{BASE}/Unbranded_Gap_Analysis.xlsx'
wb.save(out)
print(f'\nSaved: {out}')
print(f'Sheets: {[s.title for s in wb.worksheets]}')
print(f'\nSummary:')
print(f'  Total gap items: {len(results)}')
print(f'  High-demand (10+ accts): {high_demand}')
print(f'  Have SRS equivalent: {has_srs_count}')
print(f'  Recommended to add: {add_scope}')
